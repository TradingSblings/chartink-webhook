#!/usr/bin/env python3

from flask import Flask, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import json
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError

class GoogleDriveHandler:
    def __init__(self, credentials_file="credentials.json"):
        self.credentials_file = credentials_file
        self.service = None
        self.folder_id = None
        self.initialize_drive()
    
    def initialize_drive(self):
        """Initialize Google Drive API connection"""
        try:
            if not os.path.exists(self.credentials_file):
                print(f"Warning: {self.credentials_file} not found. Google Drive disabled.")
                return False
            
            # Use service account credentials
            creds = service_account.Credentials.from_service_account_file(
                self.credentials_file,
                scopes=['https://www.googleapis.com/auth/drive.file']
            )
            
            self.service = build('drive', 'v3', credentials=creds)
            print("✅ Google Drive API initialized successfully")
            return True
        except Exception as e:
            print(f"❌ Error initializing Google Drive: {e}")
            return False
    
    def create_or_get_folder(self, folder_name="Chartink_Data"):
        """Create or get the folder ID for storing files"""
        try:
            if not self.service:
                return None
            
            # Search for existing folder
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.service.files().list(q=query, fields="files(id, name)").execute()
            folders = results.get('files', [])
            
            if folders:
                self.folder_id = folders[0]['id']
                print(f"✅ Found existing folder: {folder_name}")
                return self.folder_id
            
            # Create new folder
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder'
            }
            folder = self.service.files().create(body=file_metadata, fields='id').execute()
            self.folder_id = folder.get('id')
            print(f"✅ Created new folder: {folder_name}")
            return self.folder_id
        except Exception as e:
            print(f"❌ Error creating/getting folder: {e}")
            return None
    
    def upload_file(self, file_path, file_name=None):
        """Upload file to Google Drive"""
        try:
            if not self.service:
                return None
            
            if not file_name:
                file_name = os.path.basename(file_path)
            
            # Ensure folder exists
            if not self.folder_id:
                self.create_or_get_folder()
            
            # Check if file already exists
            query = f"name='{file_name}' and '{self.folder_id}' in parents and trashed=false"
            results = self.service.files().list(q=query, fields="files(id, name)").execute()
            existing_files = results.get('files', [])
            
            file_metadata = {
                'name': file_name,
                'parents': [self.folder_id] if self.folder_id else []
            }
            
            media = MediaFileUpload(file_path, 
                                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            if existing_files:
                # Update existing file
                file_id = existing_files[0]['id']
                file = self.service.files().update(
                    fileId=file_id,
                    media_body=media
                ).execute()
                print(f"✅ Updated file in Google Drive: {file_name}")
            else:
                # Create new file
                file = self.service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id, webViewLink'
                ).execute()
                print(f"✅ Uploaded new file to Google Drive: {file_name}")
            
            return file.get('id')
        except Exception as e:
            print(f"❌ Error uploading to Google Drive: {e}")
            return None

class ChartinkWebhookHandler:
    def __init__(self, excel_file="Chartink_Workflow.xlsx", use_google_drive=True):
        self.excel_file = excel_file
        self.use_google_drive = use_google_drive
        self.drive_handler = GoogleDriveHandler() if use_google_drive else None
        self.current_iteration = self.get_next_iteration_number()
        
    def get_next_iteration_number(self):
        try:
            # Excel file will be created locally on server
            
            if not os.path.exists(self.excel_file):
                return 1
                
            workbook = load_workbook(self.excel_file)
            iteration_sheets = [sheet for sheet in workbook.sheetnames if sheet.startswith('Iteration_')]
            
            if not iteration_sheets:
                return 1
            numbers = []
            for sheet in iteration_sheets:
                try:
                    numbers.append(int(sheet.split('_')[1]))
                except (IndexError, ValueError):
                    continue
            return max(numbers) + 1 if numbers else 1
            
        except Exception as e:
            return 1
    
    def process_chartink_alert(self, alert_data):
        try:
            print("\n" + "="*60)
            print("📨 RECEIVED CHARTINK ALERT")
            print("="*60)
            print(f"📅 Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"📊 Raw Data: {json.dumps(alert_data, indent=2)}")
            print("="*60 + "\n")
            
            stocks = self.extract_stocks_from_alert(alert_data)
            if not stocks:
                print("❌ ERROR: No stocks found in alert")
                return {"status": "error", "message": "No stocks found in alert"}
            
            print(f"✅ Extracted {len(stocks)} stocks: {', '.join(stocks)}")
            result = self.add_stocks_to_excel(stocks, alert_data)
            
            if result["status"] == "success":
                print(f"✅ Successfully processed alert - Iteration {result['iteration']}")
            
            return result
        except Exception as e:
            print(f"❌ ERROR processing alert: {str(e)}")
            return {"status": "error", "message": str(e)}
    
    def extract_stocks_from_alert(self, alert_data):
        try:
            print("🔍 Extracting stocks from alert data...")
            
            if isinstance(alert_data, dict) and "stocks" in alert_data:
                if isinstance(alert_data["stocks"], str):
                    stocks = [s.strip() for s in alert_data["stocks"].split(",")]
                    print(f"   📋 Found stocks (comma-separated): {alert_data['stocks']}")
                else:
                    stocks = alert_data["stocks"]
                    print(f"   📋 Found stocks (list): {stocks}")
            else:
                print("   ⚠️  No 'stocks' field found in alert data")
                return []
            
            cleaned_stocks = []
            for stock in stocks:
                if isinstance(stock, str):
                    clean_stock = stock.strip().upper().replace("NSE:", "").replace(".NS", "")
                    if clean_stock and len(clean_stock) <= 20:
                        cleaned_stocks.append(clean_stock)
                        print(f"   ✓ Cleaned: {stock} → {clean_stock}")
            
            print(f"   ✅ Total stocks extracted: {len(cleaned_stocks)}")
            return cleaned_stocks
        except Exception as e:
            print(f"   ❌ Error extracting stocks: {str(e)}")
            return []
    
    def add_stocks_to_excel(self, stocks, alert_data):
        try:
            current_time = datetime.now()
            trigger_prices = []
            if "trigger_prices" in alert_data and isinstance(alert_data["trigger_prices"], str):
                try:
                    trigger_prices = [float(p.strip()) for p in alert_data["trigger_prices"].split(",")]
                except ValueError:
                    trigger_prices = []
            
            stock_data = []
            for i, stock in enumerate(stocks, 1):
                trigger_price = trigger_prices[i-1] if i-1 < len(trigger_prices) else 0
                
                stock_data.append({
                    "Sr_No": i,
                    "Stock_Symbol": stock,
                    "Trigger_Price": trigger_price,
                    "Scan_Name": alert_data.get("scan_name", "Unknown"),
                    "Alert_Name": alert_data.get("alert_name", "Chartink Alert"),
                    "Triggered_At": alert_data.get("triggered_at", current_time.strftime("%H:%M %p")),
                    "Date_Added": current_time.strftime("%Y-%m-%d"),
                    "Iteration": self.current_iteration,
                    "Alert_Time": current_time.strftime("%H:%M:%S"),
                    "Scan_URL": alert_data.get("scan_url", "")
                })
            
            df = pd.DataFrame(stock_data)
            
            # Create Excel file if it doesn't exist
            if not os.path.exists(self.excel_file):
                print(f"Creating new Excel file: {self.excel_file}")
                workbook = openpyxl.Workbook()
                # Remove default sheet
                workbook.remove(workbook.active)
                # Create Master_List sheet
                master_sheet = workbook.create_sheet("Master_List")
                master_headers = ["Stock_Symbol", "First_Appearance_Date", "First_Iteration", "Total_Appearances", "Last_Updated"]
                master_sheet.append(master_headers)
                # Create TradingView_Export sheet
                tv_sheet = workbook.create_sheet("TradingView_Export")
                tv_sheet.append(["TradingView_Symbols"])
                workbook.save(self.excel_file)
            else:
                workbook = load_workbook(self.excel_file)
            sheet_name = f"Iteration_{self.current_iteration}"
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.append(list(df.columns))
            for _, row in df.iterrows():
                worksheet.append(row.tolist())
            workbook.save(self.excel_file)
            
            # Upload to Google Drive if enabled
            if self.use_google_drive and self.drive_handler and self.drive_handler.service:
                self.drive_handler.upload_file(self.excel_file)
            
            self.current_iteration += 1
            
            return {
                "status": "success",
                "message": f"Added {len(stocks)} stocks to iteration {self.current_iteration - 1}",
                "stocks": stocks,
                "iteration": self.current_iteration - 1
            }
            
        except Exception as e:
            return {"status": "error", "message": str(e)}

app = Flask(__name__)
webhook_handler = ChartinkWebhookHandler()

@app.route('/webhook', methods=['POST'])
def chartink_webhook():
    try:
        print("\n🌐 Incoming webhook request...")
        print(f"   Method: {request.method}")
        print(f"   Headers: {dict(request.headers)}")
        
        alert_data = request.get_json() or request.form.to_dict()
        
        if not alert_data:
            print("❌ No data received in request")
            return jsonify({"status": "error", "message": "No data received"}), 400
        
        result = webhook_handler.process_chartink_alert(alert_data)
        
        status_code = 200 if result["status"] == "success" else 400
        print(f"📤 Sending response: {status_code} - {result['status']}\n")
        
        return jsonify(result), status_code
    except Exception as e:
        print(f"❌ Webhook error: {str(e)}\n")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/download', methods=['GET'])
def download_excel():
    try:
        excel_file = "Chartink_Workflow.xlsx"
        if os.path.exists(excel_file):
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='Chartink_Workflow.xlsx'
            )
        else:
            return jsonify({"status": "error", "message": "Excel file not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "status": "online",
        "service": "Chartink Webhook Server",
        "endpoints": {
            "webhook": "/webhook (POST) - Receive Chartink alerts",
            "download": "/download (GET) - Download Excel file"
        }
    })

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    
    print("\n" + "="*60)
    print("🚀 CHARTINK WEBHOOK SERVER STARTING")
    print("="*60)
    print(f"📡 Server URL: http://localhost:{port}")
    print(f"📥 Webhook endpoint: http://localhost:{port}/webhook")
    print(f"📥 Download endpoint: http://localhost:{port}/download")
    print(f"📂 Excel file: {webhook_handler.excel_file}")
    print(f"☁️  Google Drive: {'Enabled ✅' if webhook_handler.use_google_drive else 'Disabled ❌'}")
    print("="*60 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=False)